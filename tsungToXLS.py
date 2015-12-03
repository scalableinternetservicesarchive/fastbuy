# -*- coding: utf-8 -*-
"""
Created on Fri Nov 20 17:06:42 2015

@author: Dongqiao Ma
"""

from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference

wb = Workbook()
source_folder = "F:/tsungtest/newtsung/sqlvsbasic/"
dest_filename = 'result.xlsx'
folder_names = ['basic-medium','basic-large','basic-xlarge','basic-r3large','basic-c3large','sql-medium','sql-large','sql-xlarge','sql-r3large','sql-c3large']
for folder_name in folder_names:
    file_name = source_folder + folder_name + "/report.html"
    soup = BeautifulSoup(open(file_name))
    categories = soup.find_all("h3");
    tables = soup.find_all("table", class_="table")
    if folder_name == 'basic-medium':
        sheet = wb.active
        sheet.title = folder_name
    else:
        sheet = wb.create_sheet(folder_name)

    for index in range(len(categories)):
        if index == 1:
            continue
        category = categories[index]
        if category.text == 'Errors':
            continue
        if index >= 3:
            table = tables[index + 1]
        else:
            table = tables[index]
        rows = table.find_all("tr")
        # Add Category Title
        sheet.append([category.text])
        # Add Table Headers
        titles = rows[0].find_all("th")
        alist = []
        for i in range(len(titles)):
            alist.append(" ".join(titles[i].text.split()))
        sheet.append(alist)
        # Add Table Rows
        for row in rows[1:]:
            cols = row.find_all("td")
            alist = []
            for col in cols:
                alist.append(col.text)
            sheet.append(alist)

# Anaylysis Sheet
sheet = wb.create_sheet('analysis', 0)
rows_name = ['Medium', 'Large', 'xLarge', 'r3Large', 'c3Large'];

# Mean Response Time
sheet.append(['Mean Response', 'Basic', 'Sql'])
for index in range(len(folder_names) / 2):
    alist = []
    alist.append(rows_name[index])
    current_sheet = wb.get_sheet_by_name(folder_names[index])
    value = float(current_sheet['F5'].value.partition(' ')[0])
    alist.append(value)
    current_sheet = wb.get_sheet_by_name(folder_names[index + 5])
    value = float(current_sheet['F5'].value.partition(' ')[0])
    alist.append(value)   
    sheet.append(alist)
    
# Highest Response Time
sheet.append(['Highest Response', 'Basic', 'Sql'])
for index in range(len(folder_names) / 2):
    alist = []
    alist.append(rows_name[index])
    current_sheet = wb.get_sheet_by_name(folder_names[index])
    value = float(current_sheet['B5'].value.partition(' ')[0])
    alist.append(value)
    current_sheet = wb.get_sheet_by_name(folder_names[index + 5])
    value = float(current_sheet['B5'].value.partition(' ')[0])
    alist.append(value)   
    sheet.append(alist)

# 503
sheet.append(['503 Ratio', 'Basic', 'Sql'])
for index in range(len(folder_names) / 2):
    alist = []
    alist.append(rows_name[index])
    current_sheet = wb.get_sheet_by_name(folder_names[index])
    total_request = float(current_sheet['D27'].value) + float(current_sheet['D28'].value) +  float(current_sheet['D29'].value)
    value = float(float(current_sheet['D29'].value) / total_request)
    alist.append(value)
    current_sheet = wb.get_sheet_by_name(folder_names[index + 5])
    total_request = float(current_sheet['D27'].value) +  float(current_sheet['D28'].value) +  float(current_sheet['D29'].value)
    value = float(float(current_sheet['D29'].value) / total_request)
    alist.append(value)   
    sheet.append(alist)
    
# Draw Mean Response Chart
response_chart = BarChart()
response_chart.type = "col"
response_chart.style = 10
response_chart.title = "Mean Response Time"

data = Reference(sheet, min_col=2, min_row=1, max_row=6, max_col=3)
cats = Reference(sheet, min_col=1, min_row=2, max_row=6)
response_chart.add_data(data, titles_from_data=True)
response_chart.set_categories(cats)
response_chart.shape = 4
sheet.add_chart(response_chart, "G1")

# Draw Highest Response Chart
response_chart = BarChart()
response_chart.type = "col"
response_chart.style = 10
response_chart.title = "Highest Response Time"

data = Reference(sheet, min_col=2, min_row=7, max_row=12, max_col=3)
cats = Reference(sheet, min_col=1, min_row=8, max_row=12)
response_chart.add_data(data, titles_from_data=True)
response_chart.set_categories(cats)
response_chart.shape = 4
sheet.add_chart(response_chart, "G16")

# Draw 503 Chart
response_chart = BarChart()
response_chart.type = "col"
response_chart.style = 10
response_chart.title = "503 Ratio"

data = Reference(sheet, min_col=2, min_row=13, max_row=18, max_col=3)
cats = Reference(sheet, min_col=1, min_row=14, max_row=18)
response_chart.add_data(data, titles_from_data=True)
response_chart.set_categories(cats)
response_chart.shape = 4
sheet.add_chart(response_chart, "G32")

wb.save(dest_filename)
